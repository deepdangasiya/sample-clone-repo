import git
import os
import openpyxl
import re
import argparse


def git_readme_update(readme_data, branch):
    repo = git.Repo(os.getcwd())
    print(f"repo::: {repo} -- {branch}")
    repo.git.checkout(branch)
    print(repo.active_branch)


    # content = ("### Test Results: ###\n\n"
    #            "| Test-ID | Test Case Name | Test Case Result |\n"
    #            "| :------:|:--------------:|:----------------:|\n")
    # for ele in readme_data:
    #     content += f"|{ele['test_id']}|{ele['test_name']}|{ele['result']}|\n"
    # with open("README.md", "w") as f:
    #     f.write(content)
    # repo.index.add("README.md")
    # repo.index.commit("Update README.md with tests results")
    # # origin = repo.remote(name="origin")
    # # origin.push(branch)
    # repo.git.push("-u", "origin", branch)


def extract_excel_data():
    excel_file = "report_all.xlsx"
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    # header = [cell.value for row in
    #           ws.iter_rows(min_row=1, max_row=1, max_col=9) for cell in row]
    data_list = [[cell.value for cell in row] for row in
                 ws.iter_rows(min_row=2)]
    dict_list = []
    reg = r'(\w+)\[(\w+-\d+)]'
    for ele in data_list:
        test_name, test_id = re.search(reg, ele[1]).groups()
        temp_dict = {"test_id": test_id, "test_name": test_name,
                     "result": ele[3]}
        dict_list.append(temp_dict)
    return dict_list


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Enter git branch name.")
    parser.add_argument("--branch_name", type=str, help="Enter branch name "
                                                        "to update README.md")
    args = parser.parse_args()
    branch_name = args.branch_name
    print(f"branch_name: {branch_name.split('/')[1]}")
    data = extract_excel_data()
    git_readme_update(data, branch_name.split('/')[1])
